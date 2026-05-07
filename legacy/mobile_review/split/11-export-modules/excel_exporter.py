# excel_exporter.py - Excel export module (updated)
from pathlib import Path
from typing import Optional, List
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """مصدّر للنصوص إلى Excel (للجداول)."""

    def __init__(self, language: str = "ar"):
        self.language = language
        self.is_rtl = language == "ar"

    def export(
        self,
        html_path: Path,
        output_path: Path
    ) -> Path:
        """
        تصدير HTML إلى Excel.

        Args:
            html_path: مسار ملف HTML.
            output_path: مسار ملف Excel المخرج.

        Returns:
            Path: مسار ملف Excel.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # قراءة HTML
        with open(html_path, "r", encoding="utf-8") as f:
            html_content = f.read()

        # تحليل HTML
        soup = BeautifulSoup(html_content, "html.parser")

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "النص"

        # معالج الجداول
        tables = soup.find_all("table")
        if not tables:
            # إذا لم يتم العثور على جداول، إضافة النص كخلية واحدة
            body = soup.find("body")
            if body:
                ws.append([body.get_text()])
        else:
            for i, table in enumerate(tables):
                if i > 0:
                    ws = wb.create_sheet(title=f"جدول {i+1}")

                # إضافة رأس الجدول
                headers = [th.text for th in table.find_all("th")]
                if headers:
                    ws.append(headers)

                # إضافة البيانات
                for row in table.find_all("tr")[1:]:  # تجاهل رأس الجدول
                    cells = [td.text for td in row.find_all("td")]
                    if cells:
                        ws.append(cells)

                # تنسيق الجدول
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(
                            horizontal='right' if self.is_rtl else 'left',
                            vertical='top',
                            wrap_text=True
                        )
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

        # حفظ الملف
        wb.save(str(output_path))
        logger.info(f"تم تصدير الملف إلى Excel: {output_path}")
        return output_path
